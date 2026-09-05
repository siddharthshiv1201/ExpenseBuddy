from decimal import Decimal

from django.contrib.auth import get_user_model
from django.test import TestCase
from openpyxl import load_workbook

from accounts.models import Profile
from expenses.models import Expense

from .services import generate_monthly_expense_report


User = get_user_model()


class MonthlyReportTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(
            email="employee@example.com",
            password="TestPassword123!",
        )

        Profile.objects.create(
            user=self.user,
            first_name="Test",
            last_name="Employee",
            designation="Employee",
            state="Delhi",
            headquarters="New Delhi",
        )

    def test_multiple_expenses_same_date_are_consolidated(self):
        Expense.objects.create(
            user=self.user,
            expense_date="2026-08-25",
            category="TRAVEL",
            description="Flight",
            from_location="Delhi",
            to_location="Mumbai",
            mode_of_conveyance="Flight",
            fare=Decimal("5500.00"),
        )

        Expense.objects.create(
            user=self.user,
            expense_date="2026-08-25",
            category="MEALS",
            description="Lunch",
            food=Decimal("800.00"),
        )

        Expense.objects.create(
            user=self.user,
            expense_date="2026-08-25",
            category="TRANSPORT",
            description="Uber",
            from_location="Mumbai",
            to_location="Hotel",
            mode_of_conveyance="Uber",
            fare=Decimal("300.00"),
        )

        path = generate_monthly_expense_report(
            self.user,
            2026,
            8,
        )

        workbook = load_workbook(
            path,
            data_only=False,
        )
        worksheet = workbook.active

        # August 1 is the first calendar day.
        self.assertEqual(
            worksheet["A14"].value.strftime("%Y-%m-%d"),
            "2026-08-01",
        )

        # August 25 is row 38.
        self.assertEqual(
            worksheet["A38"].value.strftime("%Y-%m-%d"),
            "2026-08-25",
        )

        # All three expenses on August 25
        # must appear as ONE row.
        self.assertEqual(
            worksheet["H38"].value,
            Decimal("5800.00"),
        )

        self.assertEqual(
            worksheet["J38"].value,
            Decimal("800.00"),
        )

        # The next day still has its calendar date.
        self.assertEqual(
            worksheet["A39"].value.strftime("%Y-%m-%d"),
            "2026-08-26",
        )

        # Row 45 is the totals row, not another date.
        self.assertIsNone(
            worksheet["A45"].value,
        )

    def test_report_contains_employee_information(self):
        path = generate_monthly_expense_report(
            self.user,
            2026,
            8,
        )

        workbook = load_workbook(
            path,
            data_only=False,
        )
        worksheet = workbook.active

        self.assertEqual(
            worksheet["C8"].value,
            "Test Employee",
        )

        self.assertEqual(
            worksheet["E8"].value,
            "Employee",
        )

        self.assertEqual(
            worksheet["I8"].value,
            "Delhi",
        )

        self.assertEqual(
            worksheet["C9"].value,
            "New Delhi",
        )

        self.assertEqual(
            worksheet["D9"].value,
            "Month :- 2026-08",
        )

    def test_report_date_is_in_date_column(self):
        Expense.objects.create(
            user=self.user,
            expense_date="2026-08-25",
            category="TRAVEL",
            description="Flight",
            fare=Decimal("5500.00"),
        )

        path = generate_monthly_expense_report(
            self.user,
            2026,
            8,
        )

        workbook = load_workbook(
            path,
            data_only=False,
        )
        worksheet = workbook.active

        # Every day is represented in column A.
        self.assertEqual(
            worksheet["A14"].value.strftime("%Y-%m-%d"),
            "2026-08-01",
        )

        self.assertEqual(
            worksheet["A38"].value.strftime("%Y-%m-%d"),
            "2026-08-25",
        )

        self.assertEqual(
            worksheet["A14"].number_format,
            "DD/MM/YY",
        )

        # Month is shown in the header,
        # not as a separate date.
        self.assertIsNone(
            worksheet["M9"].value,
        )

    def test_individual_miscellaneous_expenses_are_reported(self):
        Expense.objects.create(
            user=self.user,
            expense_date="2026-08-25",
            category="OFFICE",
            description="Office expenses",
            phone=Decimal("100.00"),
            mobile=Decimal("200.00"),
            postage=Decimal("50.00"),
            fax=Decimal("25.00"),
            email_expense=Decimal("30.00"),
            stationary=Decimal("150.00"),
            telegram=Decimal("10.00"),
            photo_copies=Decimal("40.00"),
            octroi=Decimal("20.00"),
            demurrage=Decimal("15.00"),
            collie_cartage=Decimal("60.00"),
        )

        path = generate_monthly_expense_report(
            self.user,
            2026,
            8,
        )

        workbook = load_workbook(
            path,
            data_only=False,
        )
        worksheet = workbook.active

        # August 25 = row 38.
        #
        # Total:
        # 100 + 200 + 50 + 25 + 30
        # + 150 + 10 + 40 + 20 + 15 + 60
        # = 700
        self.assertEqual(
            worksheet["L38"].value,
            Decimal("700.00"),
        )

        # -----------------------------
        # Fixed expenses
        # -----------------------------
        self.assertEqual(
            worksheet["F49"].value,
            Decimal("100.00"),
        )

        self.assertEqual(
            worksheet["F50"].value,
            Decimal("200.00"),
        )

        self.assertEqual(
            worksheet["F51"].value,
            Decimal("50.00"),
        )

        self.assertEqual(
            worksheet["F52"].value,
            Decimal("30.00"),
        )

        # F53 contains the Excel formula.
        self.assertEqual(
            worksheet["F53"].value,
            "=F49+F50+F51+F52",
        )

        # -----------------------------
        # Other reimbursements
        # -----------------------------
        self.assertEqual(
            worksheet["H49"].value,
            Decimal("150.00"),
        )

        self.assertEqual(
            worksheet["H50"].value,
            Decimal("10.00"),
        )

        self.assertEqual(
            worksheet["H51"].value,
            Decimal("25.00"),
        )

        self.assertEqual(
            worksheet["H52"].value,
            Decimal("40.00"),
        )

        self.assertEqual(
            worksheet["M50"].value,
            Decimal("20.00"),
        )

        self.assertEqual(
            worksheet["M51"].value,
            Decimal("15.00"),
        )

        self.assertEqual(
            worksheet["M52"].value,
            Decimal("60.00"),
        )