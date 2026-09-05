from calendar import monthrange
from pathlib import Path
from tempfile import NamedTemporaryFile

from django.conf import settings
from openpyxl import load_workbook


TEMPLATE_PATH = (
    Path(settings.BASE_DIR)
    / "templates"
    / "excel"
    / "expense_report_template.xlsx"
)


MISC_FIELDS = [
    "phone",
    "mobile",
    "postage",
    "fax",
    "email_expense",
    "stationary",
    "telegram",
    "photo_copies",
    "octroi",
    "demurrage",
    "collie_cartage",
]


def amount(value):
    return value or 0


def generate_monthly_expense_report(user, year, month):
    workbook = load_workbook(TEMPLATE_PATH)
    worksheet = workbook.active

    profile = user.profile

    # -----------------------------
    # Employee / report information
    # -----------------------------
    worksheet["D7"] = profile.reporting_to
    worksheet["I7"] = profile.designation

    worksheet["C8"] = (
        f"{profile.first_name} {profile.last_name}".strip()
    )
    worksheet["E8"] = profile.designation
    worksheet["I8"] = profile.state

    worksheet["C9"] = profile.headquarters
    worksheet["D9"] = f"Month :- {year}-{month:02d}"

    # Date is displayed in the expense table.
    worksheet["M9"] = None

    # -----------------------------
    # Get expenses for the month
    # -----------------------------
    expenses = user.expenses.filter(
        expense_date__year=year,
        expense_date__month=month,
    ).order_by(
        "expense_date",
        "created_at",
    )

    expenses_by_date = {}

    for expense in expenses:
        expenses_by_date.setdefault(
            expense.expense_date,
            [],
        ).append(expense)

    # -----------------------------
    # Clear existing sample rows
    # -----------------------------
    start_row = 14
    end_row = 44

    for row in range(start_row, end_row + 1):
        for column in range(1, 16):
            worksheet.cell(
                row=row,
                column=column,
            ).value = None

    # -----------------------------
    # Create one row for every day
    # in the selected month
    # -----------------------------
    days_in_month = monthrange(year, month)[1]

    for day in range(1, days_in_month + 1):
        row_number = start_row + day - 1

        # -----------------------------
        # Date
        # -----------------------------
        from datetime import date

        expense_date = date(year, month, day)

        date_cell = worksheet.cell(
            row=row_number,
            column=1,
        )
        date_cell.value = expense_date
        date_cell.number_format = "DD/MM/YY"

        daily_expenses = expenses_by_date.get(
            expense_date,
            [],
        )

        if not daily_expenses:
            continue

        # -----------------------------
        # Descriptive information
        # -----------------------------
        from_locations = []
        to_locations = []
        modes = []
        remarks = []

        for expense in daily_expenses:
            if expense.from_location:
                from_locations.append(
                    expense.from_location
                )

            if expense.to_location:
                to_locations.append(
                    expense.to_location
                )

            if expense.mode_of_conveyance:
                modes.append(
                    expense.mode_of_conveyance
                )

            if expense.remarks:
                remarks.append(
                    expense.remarks
                )

        worksheet.cell(
            row=row_number,
            column=2,
        ).value = " / ".join(
            dict.fromkeys(from_locations)
        )

        worksheet.cell(
            row=row_number,
            column=3,
        ).value = " / ".join(
            dict.fromkeys(to_locations)
        )

        worksheet.cell(
            row=row_number,
            column=5,
        ).value = " / ".join(
            dict.fromkeys(modes)
        )

        worksheet.cell(
            row=row_number,
            column=13,
        ).value = " ; ".join(
            dict.fromkeys(remarks)
        )

        # -----------------------------
        # Distance
        # -----------------------------
        distances = [
            expense.distance_km
            for expense in daily_expenses
            if expense.distance_km is not None
        ]

        if distances:
            worksheet.cell(
                row=row_number,
                column=4,
            ).value = sum(distances)

        # -----------------------------
        # Departure / arrival times
        # -----------------------------
        departure_times = [
            expense.departure_time
            for expense in daily_expenses
            if expense.departure_time
        ]

        arrival_times = [
            expense.arrival_time
            for expense in daily_expenses
            if expense.arrival_time
        ]

        if departure_times:
            worksheet.cell(
                row=row_number,
                column=6,
            ).value = min(departure_times)

        if arrival_times:
            worksheet.cell(
                row=row_number,
                column=7,
            ).value = max(arrival_times)

        # -----------------------------
        # Main expense amounts
        # -----------------------------
        worksheet.cell(
            row=row_number,
            column=8,
        ).value = sum(
            amount(expense.fare)
            for expense in daily_expenses
        )

        worksheet.cell(
            row=row_number,
            column=9,
        ).value = sum(
            amount(expense.stay)
            for expense in daily_expenses
        )

        worksheet.cell(
            row=row_number,
            column=10,
        ).value = sum(
            amount(expense.food)
            for expense in daily_expenses
        )

        worksheet.cell(
            row=row_number,
            column=11,
        ).value = sum(
            amount(expense.da)
            for expense in daily_expenses
        )

        # -----------------------------
        # Miscellaneous total
        #
        # L = total of all the individual
        # reimbursement fields.
        #
        # miscellaneous_1 and
        # miscellaneous_2 are intentionally
        # NOT included.
        # -----------------------------
        miscellaneous_total = sum(
            amount(getattr(expense, field))
            for expense in daily_expenses
            for field in MISC_FIELDS
        )

        worksheet.cell(
            row=row_number,
            column=12,
        ).value = miscellaneous_total

    # -----------------------------
    # Monthly totals
    # -----------------------------
    worksheet["H45"] = "=SUM(H14:H44)"
    worksheet["I45"] = "=SUM(I14:I44)"
    worksheet["J45"] = "=SUM(J14:J44)"
    worksheet["K45"] = "=SUM(K14:K44)"
    worksheet["L45"] = "=SUM(L14:L44)"

    worksheet["C47"] = "=H45+I45+J45+K45+L45"

    # -----------------------------
    # Individual miscellaneous
    # expense totals
    # -----------------------------
    monthly_totals = {}

    for field in MISC_FIELDS:
        monthly_totals[field] = sum(
            amount(getattr(expense, field))
            for expense in expenses
        )

    # Fixed expenses
    worksheet["F49"] = monthly_totals["phone"]
    worksheet["F50"] = monthly_totals["mobile"]
    worksheet["F51"] = monthly_totals["postage"]
    worksheet["F52"] = monthly_totals["email_expense"]

    worksheet["F53"] = (
        "=F49+F50+F51+F52"
    )

    # Other reimbursements
   # Other reimbursements
    worksheet["H49"] = monthly_totals["stationary"]
    worksheet["H50"] = monthly_totals["telegram"]
    worksheet["H51"] = monthly_totals["fax"]
    worksheet["H52"] = monthly_totals["photo_copies"]

    worksheet["M50"] = monthly_totals["octroi"]
    worksheet["M51"] = monthly_totals["demurrage"]
    worksheet["M52"] = monthly_totals["collie_cartage"]

    # -----------------------------
    # Ensure old miscellaneous
    # detail values are blank.
    # -----------------------------
    # worksheet["N49"] = None
    # worksheet["N50"] = None
    # worksheet["N51"] = None
    # worksheet["N52"] = None
    # worksheet["N53"] = None

    # -----------------------------
    # Save generated workbook
    # -----------------------------
    temporary_file = NamedTemporaryFile(
        suffix=".xlsx",
        delete=False,
    )
    temporary_file.close()

    workbook.save(temporary_file.name)

    return temporary_file.name